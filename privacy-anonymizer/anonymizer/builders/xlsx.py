"""Reconstructor de hojas de cálculo Excel (.xlsx) anonimizadas."""

from io import BytesIO
from typing import List

from openpyxl import load_workbook

from anonymizer.builders.base import BaseBuilder
from anonymizer.extractors.base import ExtractedText


class XlsxBuilder(BaseBuilder):
    """Reconstruye un archivo Excel reemplazando el texto por su versión anonimizada."""

    def build(
        self,
        original_stream: BytesIO,
        extracted_texts: List[ExtractedText],
        redacted_texts: List[str],
    ) -> BytesIO:
        original_stream.seek(0)
        workbook = load_workbook(original_stream)

        idx = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if idx < len(extracted_texts):
                        loc = extracted_texts[idx].location
                        if (
                            loc[0] == "cell"
                            and loc[1] == sheet_name
                            and loc[2] == cell.coordinate
                        ):
                            cell.value = redacted_texts[idx]
                            idx += 1

                    if cell.comment and idx < len(extracted_texts):
                        loc = extracted_texts[idx].location
                        if (
                            loc[0] == "comment"
                            and loc[1] == sheet_name
                            and loc[2] == cell.coordinate
                        ):
                            cell.comment.text = redacted_texts[idx]
                            idx += 1

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

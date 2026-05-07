"""Extractor de texto para hojas de cálculo Excel (.xlsx)."""

from io import BytesIO
from typing import List

from openpyxl import load_workbook

from anonymizer.extractors.base import BaseExtractor, ExtractedText


class XlsxExtractor(BaseExtractor):
    """Extrae texto de archivos Excel, incluyendo celdas y comentarios."""

    def extract(self, file_stream: BytesIO) -> List[ExtractedText]:
        workbook = load_workbook(file_stream, data_only=True)
        extracted: List[ExtractedText] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        extracted.append(
                            ExtractedText(
                                text=cell.value,
                                location=("cell", sheet_name, cell.coordinate),
                            )
                        )
                    # Comentarios
                    if cell.comment and cell.comment.text:
                        extracted.append(
                            ExtractedText(
                                text=cell.comment.text,
                                location=("comment", sheet_name, cell.coordinate),
                            )
                        )

        return extracted
